#!/usr/bin/env python3
"""
AWS RDS Inventory Collector with Database Storage

Collects RDS instance information across multiple AWS regions and accounts,
and stores the data in a MySQL database for centralized inventory management.
Supports multiple AWS accounts with Business Unit (BU) tracking.
"""

import boto3
import argparse
import logging
import mysql.connector
import os
import json
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from botocore.exceptions import ClientError, NoCredentialsError

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_env_file(env_path: str = '.env'):
    """Load environment variables from .env file."""
    if os.path.exists(env_path):
        logger.info(f"📄 Loading environment variables from {env_path}")
        with open(env_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    if '=' in line:
                        key, value = line.split('=', 1)
                        os.environ[key.strip()] = value.strip()
                        logger.debug(f"  Loaded: {key.strip()}")


class RDSInventoryCollector:
    """Collects RDS inventory across multiple AWS regions and accounts."""
    
    def __init__(self, profile: str = None, regions: List[str] = None, store_in_db: bool = False, 
                 aws_account_id: str = None, bu_name: str = None, db_host: str = None):
        """
        Initialize the RDS Inventory Collector.
        
        Args:
            profile: AWS profile name to use
            regions: List of AWS regions to scan. If None, scans all available regions.
            store_in_db: Whether to store results in database
            aws_account_id: AWS Account ID for multi-account tracking
            bu_name: Business Unit name for multi-account tracking
            db_host: Database host (overrides environment variable)
        """
        self.profile = profile
        self.session = boto3.Session(profile_name=profile) if profile else boto3.Session()
        self.regions = regions or self._get_all_regions()
        self.inventory_data = []
        self.store_in_db = store_in_db
        self.db_connection = None
        self.aws_account_id = aws_account_id
        self.bu_name = bu_name or 'Default'
        self.db_host = db_host
        
        # Try to get AWS Account ID if not provided
        if not self.aws_account_id:
            self.aws_account_id = self._get_aws_account_id()
        
        if self.store_in_db:
            self._connect_to_database()
            self._create_table()
        
    def _get_aws_account_id(self) -> str:
        """Get AWS Account ID from STS."""
        try:
            sts_client = self.session.client('sts')
            account_id = sts_client.get_caller_identity()['Account']
            logger.info(f"📌 AWS Account ID: {account_id}")
            return account_id
        except Exception as e:
            logger.warning(f"Could not retrieve AWS Account ID: {e}")
            return 'UNKNOWN'
    
    def _get_all_regions(self) -> List[str]:
        """Get all available AWS regions for RDS."""
        try:
            ec2_client = self.session.client('ec2', region_name='us-east-1')
            regions = ec2_client.describe_regions()
            return [region['RegionName'] for region in regions['Regions']]
        except Exception as e:
            logger.warning(f"Could not retrieve all regions: {e}. Using default regions.")
            return ['us-east-1', 'us-west-2', 'eu-west-1', 'ap-southeast-1']
    
    def _connect_to_database(self):
        """Connect to MySQL database."""
        try:
            # Use provided db_host or get from environment
            db_host = self.db_host or os.getenv('DB_HOST')
            db_user = os.getenv('DB_USER')
            db_password = os.getenv('DB_PASSWORD')
            db_name = os.getenv('DB_NAME')
            
            if not all([db_host, db_user, db_password, db_name]):
                logger.warning("Database credentials not found. Set environment variables or create .env file:")
                logger.warning("  DB_HOST, DB_USER, DB_PASSWORD, DB_NAME")
                logger.warning("")
                logger.warning("Example .env file:")
                logger.warning("  DB_HOST=ops-prod-ue1-operations-rds-01.cetqu8suvjjy.us-east-1.rds.amazonaws.com")
                logger.warning("  DB_USER=inventory")
                logger.warning("  DB_PASSWORD=your_password")
                logger.warning("  DB_NAME=inventory")
                self.store_in_db = False
                return
            
            logger.info(f"🔗 Attempting to connect to database...")
            logger.info(f"   Host: {db_host}")
            logger.info(f"   User: {db_user}")
            logger.info(f"   Database: {db_name}")
            
            self.db_connection = mysql.connector.connect(
                host=db_host,
                user=db_user,
                password=db_password,
                database=db_name,
                autocommit=True,
                connection_timeout=30
            )
            logger.info(f"✅ Successfully connected to database: {db_name}")
        except mysql.connector.Error as e:
            logger.error(f"❌ Failed to connect to database: {e}")
            logger.error("")
            logger.error("💡 Troubleshooting:")
            logger.error("   1. Verify DB_HOST is correct (should be AWS RDS endpoint, not IP)")
            logger.error("   2. Check DB_USER and DB_PASSWORD are correct")
            logger.error("   3. Ensure your IP address is in the RDS security group inbound rules")
            logger.error("   4. Verify the database is publicly accessible")
            logger.error("   5. Test connection manually:")
            logger.error(f"      mysql -h {db_host} -u {db_user} -p")
            logger.error("")
            self.store_in_db = False
    
    def _create_table(self):
        """Create the RDS inventory table if it doesn't exist."""
        if not self.db_connection:
            return
        
        try:
            cursor = self.db_connection.cursor()
            
            create_table_query = """
            CREATE TABLE IF NOT EXISTS rds_inventory (
                id INT AUTO_INCREMENT PRIMARY KEY,
                aws_account_id VARCHAR(50),
                bu_name VARCHAR(100),
                collection_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                region VARCHAR(50),
                db_instance_identifier VARCHAR(255),
                db_instance_class VARCHAR(100),
                engine VARCHAR(50),
                engine_version VARCHAR(50),
                db_instance_status VARCHAR(50),
                master_username VARCHAR(255),
                endpoint_address VARCHAR(255),
                endpoint_port INT,
                allocated_storage INT,
                storage_type VARCHAR(50),
                multi_az BOOLEAN,
                availability_zone VARCHAR(100),
                vpc_id VARCHAR(100),
                publicly_accessible BOOLEAN,
                backup_retention_period INT,
                preferred_backup_window VARCHAR(100),
                preferred_maintenance_window VARCHAR(100),
                auto_minor_version_upgrade BOOLEAN,
                license_model VARCHAR(100),
                instance_create_time DATETIME,
                db_security_groups LONGTEXT,
                parameter_group_name VARCHAR(255),
                db_subnet_group_name VARCHAR(255),
                iops INT,
                tags LONGTEXT,
                INDEX idx_aws_account (aws_account_id),
                INDEX idx_bu_name (bu_name),
                INDEX idx_region (region),
                INDEX idx_identifier (db_instance_identifier),
                INDEX idx_timestamp (collection_timestamp)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            
            cursor.execute(create_table_query)
            logger.info("✅ Table 'rds_inventory' is ready")
            logger.info("📊 Columns: 31 fields (including AWS Account & BU tracking)")
            cursor.close()
        except mysql.connector.Error as e:
            logger.error(f"❌ Error creating table: {e}")
    
    def _insert_into_database(self, instance_data: Dict[str, Any]):
        """Insert RDS instance data in database."""
        if not self.db_connection:
            return
        
        try:
            cursor = self.db_connection.cursor()
            
            insert_query = """
            INSERT INTO rds_inventory (
                aws_account_id, bu_name, region, db_instance_identifier, db_instance_class,
                engine, engine_version, db_instance_status, master_username,
                endpoint_address, endpoint_port, allocated_storage, storage_type,
                multi_az, availability_zone, vpc_id, publicly_accessible,
                backup_retention_period, preferred_backup_window,
                preferred_maintenance_window, auto_minor_version_upgrade, license_model,
                instance_create_time, db_security_groups, parameter_group_name,
                db_subnet_group_name, iops, tags
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            """
            
            values = (
                self.aws_account_id,
                self.bu_name,
                instance_data.get('Region'),
                instance_data.get('DBInstanceIdentifier'),
                instance_data.get('DBInstanceClass'),
                instance_data.get('Engine'),
                instance_data.get('EngineVersion'),
                instance_data.get('DBInstanceStatus'),
                instance_data.get('MasterUsername'),
                instance_data.get('Endpoint'),
                instance_data.get('Port'),
                instance_data.get('AllocatedStorage'),
                instance_data.get('StorageType'),
                instance_data.get('MultiAZ'),
                instance_data.get('AvailabilityZone'),
                instance_data.get('VpcId'),
                instance_data.get('PubliclyAccessible'),
                instance_data.get('BackupRetentionPeriod'),
                instance_data.get('PreferredBackupWindow'),
                instance_data.get('PreferredMaintenanceWindow'),
                instance_data.get('AutoMinorVersionUpgrade'),
                instance_data.get('LicenseModel'),
                instance_data.get('InstanceCreateTime'),
                instance_data.get('DBSecurityGroups'),
                instance_data.get('ParameterGroupName'),
                instance_data.get('DBSubnetGroupName'),
                instance_data.get('Iops'),
                instance_data.get('Tags')
            )
            
            cursor.execute(insert_query, values)
            logger.debug(f"Inserted: {instance_data.get('DBInstanceIdentifier')} (BU: {self.bu_name})")
            cursor.close()
        except mysql.connector.Error as e:
            logger.error(f"❌ Error inserting data into database: {e}")
    
    def collect_rds_instances(self) -> List[Dict[str, Any]]:
        """Collect RDS instances from all specified regions."""
        logger.info(f"🔍 Starting RDS inventory collection for {self.bu_name} (Account: {self.aws_account_id})")
        logger.info(f"📍 Scanning {len(self.regions)} regions")
        
        for region in self.regions:
            logger.info(f"📍 Scanning region: {region}")
            try:
                rds_client = self.session.client('rds', region_name=region)
                paginator = rds_client.get_paginator('describe_db_instances')
                
                for page in paginator.paginate():
                    for db_instance in page['DBInstances']:
                        instance_data = self._extract_instance_data(db_instance, region)
                        self.inventory_data.append(instance_data)
                        
                        # Insert into database if enabled
                        if self.store_in_db:
                            self._insert_into_database(instance_data)
                        
                        logger.debug(f"Found instance: {instance_data['DBInstanceIdentifier']}")
                        
            except ClientError as e:
                logger.error(f"❌ Error scanning region {region}: {e}")
            except Exception as e:
                logger.error(f"❌ Unexpected error in region {region}: {e}")
        
        logger.info(f"✅ Collection complete for {self.bu_name}. Found {len(self.inventory_data)} RDS instances.")
        return self.inventory_data
    
    def _extract_instance_data(self, db_instance: Dict, region: str) -> Dict[str, Any]:
        """Extract relevant data from RDS instance description."""
        
        # Extract security groups
        db_security_groups = []
        if db_instance.get('DBSecurityGroups'):
            db_security_groups = [sg.get('DBSecurityGroupName', 'N/A') for sg in db_instance.get('DBSecurityGroups', [])]
        db_security_groups_str = json.dumps(db_security_groups) if db_security_groups else 'N/A'
        
        # Extract parameter group name
        parameter_group_name = 'N/A'
        if db_instance.get('DBParameterGroups'):
            parameter_group_name = db_instance.get('DBParameterGroups', [{}])[0].get('DBParameterGroupName', 'N/A')
        
        # Extract subnet group name
        db_subnet_group_name = 'N/A'
        if db_instance.get('DBSubnetGroup'):
            db_subnet_group_name = db_instance.get('DBSubnetGroup', {}).get('DBSubnetGroupName', 'N/A')
        
        # Extract IOPS
        iops = db_instance.get('Iops', None)
        
        # Extract tags
        tags = {}
        if db_instance.get('TagList'):
            tags = {tag.get('Key', ''): tag.get('Value', '') for tag in db_instance.get('TagList', [])}
        tags_str = json.dumps(tags) if tags else 'N/A'
        
        return {
            'Region': region,
            'DBInstanceIdentifier': db_instance.get('DBInstanceIdentifier', 'N/A'),
            'DBInstanceClass': db_instance.get('DBInstanceClass', 'N/A'),
            'Engine': db_instance.get('Engine', 'N/A'),
            'EngineVersion': db_instance.get('EngineVersion', 'N/A'),
            'DBInstanceStatus': db_instance.get('DBInstanceStatus', 'N/A'),
            'MasterUsername': db_instance.get('MasterUsername', 'N/A'),
            'Endpoint': db_instance.get('Endpoint', {}).get('Address', 'N/A') if db_instance.get('Endpoint') else 'N/A',
            'Port': db_instance.get('Endpoint', {}).get('Port', 'N/A') if db_instance.get('Endpoint') else 'N/A',
            'AllocatedStorage': db_instance.get('AllocatedStorage', 'N/A'),
            'StorageType': db_instance.get('StorageType', 'N/A'),
            'MultiAZ': db_instance.get('MultiAZ', False),
            'AvailabilityZone': db_instance.get('AvailabilityZone', 'N/A'),
            'VpcId': db_instance.get('DBSubnetGroup', {}).get('VpcId', 'N/A') if db_instance.get('DBSubnetGroup') else 'N/A',
            'PubliclyAccessible': db_instance.get('PubliclyAccessible', False),
            'BackupRetentionPeriod': db_instance.get('BackupRetentionPeriod', 'N/A'),
            'PreferredBackupWindow': db_instance.get('PreferredBackupWindow', 'N/A'),
            'PreferredMaintenanceWindow': db_instance.get('PreferredMaintenanceWindow', 'N/A'),
            'AutoMinorVersionUpgrade': db_instance.get('AutoMinorVersionUpgrade', False),
            'LicenseModel': db_instance.get('LicenseModel', 'N/A'),
            'InstanceCreateTime': db_instance.get('InstanceCreateTime', 'N/A'),
            'DBSecurityGroups': db_security_groups_str,
            'ParameterGroupName': parameter_group_name,
            'DBSubnetGroupName': db_subnet_group_name,
            'Iops': iops,
            'Tags': tags_str,
        }
    
    def export_to_excel(self, filename: str = None):
        """Export collected inventory to Excel file."""
        if not self.inventory_data:
            logger.warning("No inventory data to export.")
            return
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"rds_inventory_{self.bu_name}_{timestamp}.xlsx"
        
        logger.info(f"📊 Exporting inventory to {filename}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "RDS Inventory"
        
        # Define headers
        headers = list(self.inventory_data[0].keys())
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, instance in enumerate(self.inventory_data, 2):
            for col_num, header in enumerate(headers, 1):
                value = instance.get(header, 'N/A')
                # Convert datetime objects to strings
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        wb.save(filename)
        logger.info(f"✅ Export complete: {filename}")
    
    def close_database(self):
        """Close database connection."""
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
            logger.info("Database connection closed")


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Collect AWS RDS inventory across multiple regions, accounts, and BUs'
    )
    parser.add_argument(
        '--profile',
        help='AWS profile name to use (default: default profile)',
        default=None
    )
    parser.add_argument(
        '--regions',
        nargs='+',
        help='Specific AWS regions to scan (default: all regions)',
        default=None
    )
    parser.add_argument(
        '--output',
        help='Output Excel filename (default: rds_inventory_BU_TIMESTAMP.xlsx)',
        default=None
    )
    parser.add_argument(
        '--db',
        action='store_true',
        help='Store inventory in MySQL database (requires DB_* environment variables or .env file)'
    )
    parser.add_argument(
        '--account-id',
        help='AWS Account ID (for multi-account tracking)',
        default=None
    )
    parser.add_argument(
        '--bu-name',
        help='Business Unit name (for multi-account tracking)',
        default=None
    )
    parser.add_argument(
        '--db-host',
        help='Database host override (default: use DB_HOST env variable or .env file)',
        default=None
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Enable debug logging'
    )
    
    args = parser.parse_args()
    
    if args.debug:
        logger.setLevel(logging.DEBUG)
    
    # Load .env file if it exists
    load_env_file('.env')
    
    try:
        collector = RDSInventoryCollector(
            profile=args.profile,
            regions=args.regions,
            store_in_db=args.db,
            aws_account_id=args.account_id,
            bu_name=args.bu_name,
            db_host=args.db_host
        )
        collector.collect_rds_instances()
        
        # Export to Excel if not only storing in DB
        if not args.db or args.output:
            collector.export_to_excel(filename=args.output)
        
        collector.close_database()
        
    except NoCredentialsError:
        logger.error("❌ AWS credentials not found. Please configure your credentials.")
        return 1
    except Exception as e:
        logger.error(f"❌ An error occurred: {e}")
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())
